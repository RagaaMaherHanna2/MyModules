import base64
from datetime import datetime, time

from dateutil.relativedelta import relativedelta

from odoo import models, fields, api, _
import logging

from datetime import datetime, timedelta

from odoo.addons.redeemly_pin_management.constants import DATETIME_FORMAT, DATE_FORMAT
from odoo.addons.redeemly_pin_management.controllers.pull_controller import PullController
from odoo.tools import config
import xlwt
from openpyxl import Workbook
import base64
from io import BytesIO
from odoo.addons.redeemly_pin_management.services.notification_service import NotificationService

_logger = logging.getLogger(__name__)


class SaleOrderLinePinManagement(models.Model):
    _inherit = 'sale.order.line'

    invoiced_line = fields.Boolean(string='Invoiced', index=True, default=False)
    merchant_tax = fields.Float(store=True)


class SaleOrderPinManagement(models.Model):
    _inherit = 'sale.order'

    service_provider_id = fields.Many2one('res.users', domain=[('is_service_provider', '=', True)], required=True,
                                          index=True)
    is_invoice_request = fields.Boolean(default=False)

    def action_confirm(self):
        res = super(SaleOrderPinManagement, self).action_confirm()
        return res

    def get_pulled_serials(self):
        serials = self.env['product.serials'].sudo().search([('order_id', '=', self.id)])
        aes_cipher = self.env['aes.cipher'].create([])
        pulled_serials = []
        for serial in serials:
            pulled_serials.append({
                "serial_number": serial.serial_number,
                "serial_code": aes_cipher.decrypt(serial.serial_code),
                'product_id': serial.product_id.id,
                'product_name': serial.product_id.name,
                'state': PullController.get_serial_status(serial.state),
                'SKU': serial.product_id.SKU,
                'expiry_date': datetime.strftime(serial.expiry_date, DATETIME_FORMAT) if serial.expiry_date else None,
            })
        return pulled_serials

    def get_cron_data(self, frequency):
        sps = self.env['res.users'].search([('is_service_provider', '=', True),
                                            ('enable_invoice_auto_generating', '=', True),
                                            ('invoice_generating_frequency', '=', frequency)])
        to_date = datetime.combine(datetime.now() - relativedelta(days=1), time.max)

        if frequency == 'daily':
            from_date = datetime.combine(to_date, time.min)
        elif frequency == 'weekly':
            from_date = datetime.combine(datetime.now() - relativedelta(weeks=1), time.min)
        else:
            from_date = datetime.combine(datetime.now() - relativedelta(months=1), time.min)
        return sps, from_date, to_date

    def process_invoice_request(self, service_providers, from_date, to_date):
        _logger.info("service_providers is %s", service_providers)
        _logger.info("from_date is %s", from_date)
        _logger.info("from_date is %s", to_date)
        all_orders = self.env['sale.order'].search([])
        for o in all_orders:
            _logger.info('order is %s and its date is %s', o.id, o.date_order)
        orders = self.env['sale.order'].search([('is_invoice_request', '=', False),
                                                ('date_order', '>=', from_date),
                                                ('date_order', '<=', to_date),
                                                ])
        _logger.info("orders is %s", orders)
        for sp in service_providers:
            so_ids = orders.filtered(lambda o: o.service_provider_id == sp)
            if so_ids:
                users = self.env['res.users'].search([])
                merchants = so_ids.mapped('partner_id')
                for merchant in merchants:
                    merchant_user = users.filtered(lambda u: u.partner_id == merchant) if len(
                        users.filtered(lambda u: u.partner_id == merchant)) == 1 else \
                        users.filtered(lambda u: u.partner_id == merchant)[0]
                    self.env['merchant.invoice.request'].sudo().create({
                        'service_provider_id': sp.id,
                        'merchant': merchant_user.id,
                        'from_date': from_date,
                        'to_date': to_date,
                        'type': 'system'
                    })

                so_ids.is_invoice_request = True

    def daily_process_invoice_request(self):
        try:
            service_providers, from_date, to_date = self.get_cron_data('daily')
            self.process_invoice_request(service_providers, from_date, to_date)
        except Exception as e:
            _logger.info("skarla_cron Process daily invoice create")
            _logger.exception(e)
            self.env.cr.rollback()

    def weekly_process_invoice_request(self):
        try:
            service_providers, from_date, to_date = self.get_cron_data('weekly')
            self.process_invoice_request(service_providers, from_date, to_date)
        except Exception as e:
            _logger.info("skarla_cron Process weekly invoice create")
            _logger.exception(e)
            self.env.cr.rollback()

    def monthly_process_invoice_request(self):
        try:
            service_providers, from_date, to_date = self.get_cron_data('monthly')
            self.process_invoice_request(service_providers, from_date, to_date)
        except Exception as e:
            _logger.info("skarla_cron Process monthly invoice create")
            _logger.exception(e)
            self.env.cr.rollback()


class InvoiceRequest(models.Model):
    _name = 'merchant.invoice.request'
    _inherit = ['mail.thread']

    request_date = fields.Datetime(string='Request Date', default=lambda self: fields.Datetime.today(), required=True,
                                   index=True)
    invoice_number = fields.Integer('Invoice Number', default=1)
    service_provider_id = fields.Many2one('res.users', domain=[('is_service_provider', '=', True)], index=True)
    # TODO to enhance this to be partner not user
    merchant = fields.Many2one('res.users', domain=[('is_merchant', '=', True)], index=True)
    from_date = fields.Datetime(string='From Date')
    to_date = fields.Datetime(string='To Date')
    state = fields.Selection(selection=[
        ('pending', 'Pending'),
        ('success', 'Success'),
        ('failed', 'Failed'),
    ], default='pending')
    failure_reason = fields.Char(string='Failure Reason')
    invoice = fields.Binary(string='Invoice Image')
    show_on_merchant_dashboard = fields.Boolean(string='show on merchant dashboard', default=False, index=True)
    type = fields.Selection([
        ('system', 'BY System'),
        ('sp', 'By Service Provider'),
    ]
    )

    def get_image_url(self):
        self.ensure_one()
        self = self.sudo()
        attachment = self.env['ir.attachment'].with_user(1).search([('type', '=', 'binary'),
                                                                    ('res_model', '=', self._name),
                                                                    ('res_id', '=', self.id),
                                                                    ('res_field', '=', 'invoice'),
                                                                    ('url', '!=', False)
                                                                    ], limit=1)
        print(attachment.url)
        return attachment.url if attachment else False

    def get_public_url(self):
        self.ensure_one()
        url = self.get_image_url()
        base_url = self.env['ir.config_parameter'].sudo().get_param('s3.obj_url')
        print(self.env['ir.config_parameter'].sudo().get_param('s3.obj_url'), '=========',
              url.replace(base_url, "/exposed/download_attachment?file_hash="))
        return url.replace(base_url, "/exposed/download_attachment?file_hash=")

    def get_public_url_attachment(self):
        self.ensure_one()
        url = self.get_image_url()
        base_url = self.env['ir.config_parameter'].sudo().get_param('s3.obj_url')
        return url

    def get_sp_currency(self):
        return self.service_provider_id.sp_currency.symbol

    def print_invoice_test(self):

        if self.from_date and self.from_date:
            sql = """
                           select pt.id, pt.name, sol.price_unit, SUM(sol.price_subtotal) as price_subtotal,
                            SUM(sol.product_uom_qty) as quantity_subtotal
                           from public.sale_order_line sol
                               join public.sale_order so on so.id = sol.order_id
                               join public.product_product pp on pp.id = sol.product_id
                               join public.product_template pt on pt.id = pp.product_tmpl_id
                           where so.partner_id = %s
                           and so.service_provider_id = %s
                           and so.date_order between %s and %s
                           group by pt.id, pt.name, sol.price_unit
                       """
            self.env.cr.execute(sql, (self.merchant.partner_id.id,
                                      self.service_provider_id.id,
                                      self.from_date, self.to_date))
        else:
            sql = """
                                       select pt.id, pt.name, sol.price_unit, SUM(sol.price_subtotal) as price_subtotal,
                                        SUM(sol.product_uom_qty) as quantity_subtotal
                                       from public.sale_order_line sol
                                           join public.sale_order so on so.id = sol.order_id
                                           join public.product_product pp on pp.id = sol.product_id
                                           join public.product_template pt on pt.id = pp.product_tmpl_id
                                       where so.partner_id = %s
                                       and so.service_provider_id = %s
                                       group by pt.id, pt.name, sol.price_unit
                                   """
            self.env.cr.execute(sql, (self.merchant.partner_id.id,
                                      self.service_provider_id.id))
        page_data = self.env.cr.dictfetchall()

        qr_pdf = \
            self.env.ref('redeemly_pin_management.invoice_merchant_template_report')._render_qweb_pdf(
                res_ids=self.ids,
                data={"data": page_data})[0]
        qr_pdf = base64.b64encode(qr_pdf)
        self.invoice = qr_pdf

    def process_pending_request(self, request_type):
        try:
            pending_request = self.env['merchant.invoice.request'].sudo().search(
                [('state', '=', 'pending'), ('type', '=', request_type)])
            lst_implement_sp = []
            blocked_implement_sp = []
            if request_type == 'system':
                real_count = config.get('number_invoice_request_sys')
            else:
                real_count = config.get('number_invoice_request_sp')
            for invoice_request in pending_request:
                # TODO to enhance lists filling
                if invoice_request.service_provider_id.id in blocked_implement_sp:
                    continue
                count = 0
                for item in lst_implement_sp:
                    if item == invoice_request.service_provider_id.id:
                        count = count + 1
                if count >= int(real_count):
                    blocked_implement_sp.append(invoice_request.service_provider_id.id)
                    continue

                if invoice_request.from_date and invoice_request.to_date:
                    sql = """
                        select pt.id, pt.name, sol.price_unit , sol.merchant_tax, SUM(sol.price_subtotal) as price_subtotal,
                         SUM(sol.product_uom_qty) as quantity_subtotal , SUM(sol.price_tax) as price_tax
                        from public.sale_order_line sol
                            join public.sale_order so on so.id = sol.order_id
                            join public.product_product pp on pp.id = sol.product_id
                            join public.product_template pt on pt.id = pp.product_tmpl_id
                        where so.partner_id = %s
                        and so.service_provider_id = %s
                        and so.date_order between %s and %s
                        group by pt.id, pt.name, sol.price_unit , sol.merchant_tax
                    """
                    self.env.cr.execute(sql, (invoice_request.merchant.partner_id.id,
                                              invoice_request.service_provider_id.id,
                                              invoice_request.from_date,
                                              datetime(invoice_request.to_date.year, invoice_request.to_date.month,
                                                       invoice_request.to_date.day, 23, 59, 59)))

                page_data = self.env.cr.dictfetchall()
                if not page_data:
                    invoice_request.failure_reason = "No order found for this range"
                    invoice_request.state = 'failed'
                    NotificationService.Send_Notification_email(user=invoice_request.service_provider_id,
                                                                self=invoice_request)
                    continue

                last_invoice_id = self.search([('service_provider_id', '=', invoice_request.service_provider_id.id),
                                               ('state', '=', 'success')
                                               ], order='id desc', limit=1)
                if not last_invoice_id:
                    invoice_number = 1
                else:
                    invoice_number = last_invoice_id.invoice_number + 1
                invoice_request.invoice_number = invoice_number

                qr_pdf = \
                    self.env.ref('redeemly_pin_management.invoice_merchant_template_report')._render_qweb_pdf(
                        res_ids=invoice_request.ids,
                        data={"data": page_data})[0]
                qr_pdf = base64.b64encode(qr_pdf)

                invoice_request.invoice = qr_pdf
                invoice_request.state = 'success'
                template = self.env.ref('redeemly_pin_management.merchant_invoice_request_email_tempalte')
                context = {'server_base_url': config.get('server_base_url')}
                email_to = self.merchant.merchant_invoice_mail or self.merchant.login
                email_values = {
                    'email_from': 'noreply@skarla.com',
                    'email_to': email_to
                }
                template.with_context(context).send_mail(invoice_request.id, email_values=email_values)
                NotificationService.Send_Notification_email(user=invoice_request.service_provider_id,
                                                            self=invoice_request)
                lst_implement_sp.append(invoice_request.service_provider_id.id)
        except Exception as e:
            _logger.info("skarla_cron pending request process")
            _logger.exception(e)
            self.env.cr.rollback()

    def process_pending_request_for_system(self):
        try:
            self.process_pending_request('system')
        except Exception as e:
            _logger.info("skarla_cron pending request for system")
            _logger.exception(e)
            self.env.cr.rollback()

    def process_pending_request_for_sp(self):
        try:
            self.process_pending_request('sp')
        except Exception as e:
            _logger.info("skarla_cron pending request for sp")
            _logger.exception(e)
            self.env.cr.rollback()

    def serialize_for_api(self):
        self.ensure_one()
        return {
            "id": self.id,
            "state": self.state,
            "failure_reason": self.failure_reason if self.failure_reason else "",
            'from_date': datetime.strftime(self.from_date, DATE_FORMAT) if self.from_date else None,
            'to_date': datetime.strftime(self.to_date, DATE_FORMAT) if self.to_date else None,
            'request_date': datetime.strftime(self.request_date, DATETIME_FORMAT) if self.request_date else None,
            'merchant': {
                'name': self.merchant.name,
                'reference': self.merchant.reference,
            },
            'image': self.get_public_url() if self.state == 'success' else False,
            'invoice_number': self.invoice_number,
            'show_on_merchant_dashboard': self.show_on_merchant_dashboard,
            "type": self.type

        }


class SalesReport(models.Model):
    _name = 'detail.sales.report'
    report_date = fields.Datetime()
    report_excel = fields.Binary(string='Report Excel')
    from_date = fields.Date()
    to_date = fields.Date()
    merchant_filter = fields.Many2one('res.users')
    product_id = fields.Char("Product")
    product = fields.Many2one('product.template')
    user_id = fields.Many2one('res.users')
    failure_reason = fields.Char(string="Failure Reason")
    state = fields.Selection([
        ('pending', 'Pending'),
        ('done', 'Done'),
        ('failed', 'Failed')
    ], default='pending')

    def get_image_url(self):
        self.ensure_one()
        self = self.sudo()
        attachment = self.env['ir.attachment'].with_user(1).search([('type', '=', 'binary'),
                                                                    ('res_model', '=', self._name),
                                                                    ('res_id', '=', self.id),
                                                                    ('res_field', '=', 'report_excel'),
                                                                    ('url', '!=', False)
                                                                    ], limit=1)
        return attachment.url if attachment else False

    def get_public_url(self):
        self.ensure_one()
        url = self.get_image_url()

        _logger.info('==============url===============%s', url)
        base_url = self.env['ir.config_parameter'].sudo().get_param('s3.obj_url')
        return url.replace(base_url, "/exposed/download_detail_sales_report?file_hash=")

    def get_product_name(self):
        self.ensure_one()
        if self.product:
            return self.product.name
        if self.product_id:
            return self.product_id
        return False

    def process_sales_report_request(self):
        try:
            report_requests = self.search([('state', '=', 'pending')])
            processed_sp = {}
            processed_merchant = {}
            row_count_limit = 500000
            sp_limit = merchant_limit = int(config.get('number_invoice_request_sp'))

            for rec in report_requests:
                if rec.user_id.is_service_provider:
                    # limit processed request for each sp to config limit
                    if rec.user_id.id in processed_sp:
                        processed_sp[rec.user_id.id] = processed_sp[rec.user_id.id] + 1
                    else:
                        processed_sp[rec.user_id.id] = 1
                    if processed_sp[rec.user_id.id] > sp_limit:
                        continue
                    params = [rec.from_date, datetime(rec.to_date.year, rec.to_date.month,
                                                      rec.to_date.day, 23, 59, 59), rec.user_id.id]
                    sql = """
                            select s.serial_number, p.name, p."SKU", so.date_order, so.name, so.amount_total from 
                                public.product_template p
                                join public.product_serials s on p.id = s.product_id
                                join public.sale_order so on so.id = s.order_id
                                where
                                so.date_order between %s and %s
                                and so.service_provider_id = %s 
                            """
                    if rec.get_product_name():
                        sql += " and p.name like %s"
                        params.append(rec.get_product_name())

                    if rec.merchant_filter:
                        sql += " and so.partner_id = %s"
                        params.append(rec.merchant_filter.partner_id.id)

                    self._cr.execute("select count(*) from (" + sql + ") as subinner", params)
                    sql_count = self._cr.fetchall()[0][0]
                    if sql_count > row_count_limit:
                        rec.state = 'failed'
                        rec.failure_reason = "Number Of Fetched Data Excesses Limit (%s) " \
                                             "Please Narrow Date Range" % row_count_limit
                        continue

                    self._cr.execute(sql, params)
                if rec.user_id.is_merchant:
                    params = [rec.from_date, datetime(rec.to_date.year, rec.to_date.month,
                                                      rec.to_date.day, 23, 59, 59), rec.user_id.partner_id.id]
                    if rec.user_id.id in processed_merchant:
                        processed_merchant[rec.user_id.id] = processed_merchant[rec.user_id.id] + 1
                    else:
                        processed_merchant[rec.user_id.id] = 1
                    if processed_merchant[rec.user_id.id] > merchant_limit:
                        continue

                    sql = """
                            select s.serial_number, p.name, p."SKU", so.date_order, so.name, so.amount_total from 
                                public.product_template p
                                join public.product_serials s on p.id = s.product_id
                                join public.sale_order so on so.id = s.order_id
                                where
                                so.date_order between %s and %s
                                and so.partner_id = %s 
                            """

                    if rec.get_product_name():
                        sql += " and p.name like %s"
                        params.append(rec.get_product_name())

                    self._cr.execute("select count(*) from (" + sql + ") as subinner", params)
                    sql_count = self._cr.fetchall()[0][0]
                    if sql_count > row_count_limit:
                        rec.state = 'failed'
                        rec.failure_reason = "Number Of Fetched Data Excesses Limit (%s) " \
                                             "Please Narrow Date Range" % row_count_limit
                        continue
                    self._cr.execute(sql, params)

                res = self._cr.fetchall()
                if len(res) > 0:
                    workbook = Workbook()
                    format1 = xlwt.easyxf('font:bold True;pattern: pattern solid, fore_colour gray25;align: horiz left; borders: top_color black, bottom_color black, right_color black, left_color black,\
                                                                                                         left thin, right thin, top thin, bottom thin;')
                    sheet = workbook.active
                    sheet.cell(row=1, column=1).value = 'Serial Number'
                    sheet.cell(row=1, column=2).value = 'Product'
                    sheet.cell(row=1, column=3).value = 'SKU'
                    sheet.cell(row=1, column=4).value = 'Order Date'
                    sheet.cell(row=1, column=5).value = 'Order Reference'
                    sheet.cell(row=1, column=6).value = 'Order amount'
                    index = 2
                    for item in res:
                        sheet.cell(row=index, column=1).value = item[0]
                        sheet.cell(row=index, column=2).value = item[1]
                        sheet.cell(row=index, column=3).value = item[2]
                        sheet.cell(row=index, column=4).value = item[3]
                        sheet.cell(row=index, column=5).value = item[4]
                        sheet.cell(row=index, column=6).value = item[5]
                        index = index + 1
                    fp = BytesIO()
                    workbook.save(fp)
                    rec.report_excel = base64.b64encode(fp.getvalue())
                    fp.close()
                    rec.state = 'done'
                else:
                    rec.state = 'failed'
                    rec.failure_reason = "No Data"
        except Exception as e:
            _logger.info("skarla_cron sales report request")
            _logger.exception(e)
            self.env.cr.rollback()

    def serialize_for_api(self):
        self.ensure_one()
        return {
            'id': self.id,
            'report_date': datetime.strftime(self.report_date, DATETIME_FORMAT),
            'from_date': datetime.strftime(self.from_date, DATE_FORMAT),
            'to_date': datetime.strftime(self.to_date, DATE_FORMAT),
            'failure_reason': self.failure_reason,
            'state': self.state,
            'file': self.get_public_url() if self.state == 'done' else False,
            'merchant_filter': {
                "id": self.merchant_filter.id,
                "name": self.merchant_filter.name,
            } if self.merchant_filter else {},
            "product": self.get_product_name()
        }

